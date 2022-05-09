from faker import Faker
from openpyxl import Workbook
from time import time

fake_data = Faker(['pl_PL'])

wb = Workbook()
ws = wb.active

record_count = input("Podaj liczbe wygenerowanych danych: ")
a = int(record_count)

#za duzo forow, funkcja
for i in range(1, (a + 1)):
    for j in range(1, 100):
        ws.cell(row=i, column=2).value = fake_data.first_name()
        ws.cell(row=i, column=3).value = fake_data.last_name()
        ws.cell(row=i, column=4).value = fake_data.postcode()
        ws.cell(row=i, column=5).value = fake_data.phone_number()

#funkcja
if __name__ == '__main__':
    start = time()
    wb.save("xxxx.xlsx")
    elapsed = time() - start
    print('created openpyxl file time: {}'.format(elapsed))

import argparse

parser = argparse.ArgumentParser(description="Generuj")
parser.add_argument('base', type=int, help="dane")
parser.add_argument('exponent', type=int, help="ilosc")
args = parser.parse_args()
print(args.base ** args.exponent)